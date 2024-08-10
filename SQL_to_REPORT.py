#!/usr/bin/env python
# coding: utf-8

# In[7]:


import tkinter as tk
from tkinter import ttk
import tkinter.messagebox as messagebox
from tkcalendar import DateEntry
import pandas as pd
from sqlalchemy import create_engine, text
from openpyxl import load_workbook

# Database connection information
DATABASE_NAME = "PLC_DATA"
SERVER_NAME = "SUBHAJIT\\SQLEXPRESS"
DRIVER = "ODBC Driver 17 for SQL Server"

# Connection string
connection_string = f"mssql+pyodbc://@{SERVER_NAME}/{DATABASE_NAME}?driver={DRIVER}"

# Template file paths
TEMPLATES = {
    "TAG_VALUES": "TAG_VALUES_TEMPLATE.xlsx",
    "TAG_VALUES_1": "TAG_VALUES_1_TEMPLATE.xlsx"
}

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Database Query to Excel")

        # Start Date and Time
        self.label_start_date = tk.Label(root, text="Start Date")
        self.label_start_date.grid(row=0, column=0, padx=10, pady=10)
        self.start_date = DateEntry(root, date_pattern='yyyy-mm-dd')
        self.start_date.grid(row=0, column=1, padx=10, pady=10)
        
        self.label_start_time = tk.Label(root, text="Start Time (HH:MM:SS)")
        self.label_start_time.grid(row=0, column=2, padx=10, pady=10)
        self.start_time = ttk.Combobox(root, values=[f'{str(h).zfill(2)}:00:00' for h in range(24)])
        self.start_time.grid(row=0, column=3, padx=10, pady=10)
        self.start_time.set("00:00:00")  # Default start time
        
        # End Date and Time
        self.label_end_date = tk.Label(root, text="End Date")
        self.label_end_date.grid(row=1, column=0, padx=10, pady=10)
        self.end_date = DateEntry(root, date_pattern='yyyy-mm-dd')
        self.end_date.grid(row=1, column=1, padx=10, pady=10)

        self.label_end_time = tk.Label(root, text="End Time (HH:MM:SS)")
        self.label_end_time.grid(row=1, column=2, padx=10, pady=10)
        self.end_time = ttk.Combobox(root, values=[f'{str(h).zfill(2)}:00:00' for h in range(24)])
        self.end_time.grid(row=1, column=3, padx=10, pady=10)
        self.end_time.set("23:59:59")  # Default end time
        
        # Table Dropdown
        self.label_table = tk.Label(root, text="Select Table")
        self.label_table.grid(row=2, column=0, padx=10, pady=10)
        self.table_var = tk.StringVar(root)
        self.table_dropdown = ttk.Combobox(root, textvariable=self.table_var)
        self.table_dropdown.grid(row=2, column=1, padx=10, pady=10)

        # Populate tables dropdown
        self.populate_tables()

        # Generate Button
        self.generate_button = tk.Button(root, text="Generate Excel", command=self.generate_excel)
        self.generate_button.grid(row=3, columnspan=4, pady=20)
        
        # Footer Label
        self.footer_label = tk.Label(root, text="Developed with love by Subhajit")
        self.footer_label.grid(row=4, columnspan=4, pady=10)

    def populate_tables(self):
        try:
            # Connect to the database
            engine = create_engine(connection_string)
            query = "SELECT table_name FROM information_schema.tables WHERE table_type = 'BASE TABLE'"
            with engine.connect() as connection:
                result = connection.execute(text(query))
                tables = [row['table_name'] for row in result]
                self.table_dropdown['values'] = tables
                if tables:
                    self.table_dropdown.set(tables[0])  # Select the first table by default
        except Exception as e:
            messagebox.showerror("Error", f"Could not fetch tables: {e}")

    def generate_excel(self):
        start_datetime = f"{self.start_date.get()} {self.start_time.get()}"
        end_datetime = f"{self.end_date.get()} {self.end_time.get()}"
        selected_table = self.table_var.get()

        if not start_datetime or not end_datetime or not selected_table:
            messagebox.showerror("Error", "All fields are required!")
            return

        try:
            # Load the appropriate template
            template_file = TEMPLATES.get(selected_table)
            if not template_file:
                messagebox.showerror("Error", f"No template found for table {selected_table}")
                return

            # Load the Excel template
            wb = load_workbook(template_file)
            ws = wb.active

            # Set the start and end dates in the template
            ws["B3"] = start_datetime
            ws["B4"] = end_datetime

            # Connect to the database and fetch the data
            engine = create_engine(connection_string)
            query = f"SELECT * FROM {selected_table} WHERE record_datetime BETWEEN '{start_datetime}' AND '{end_datetime}'"
            df = pd.read_sql(query, engine)

            # Write data to the template starting from row 7
            for row in df.itertuples(index=False):
                ws.append(row)

            # Save to Excel
            output_file = f"{selected_table}_output.xlsx"
            wb.save(output_file)

            messagebox.showinfo("Success", f"Excel file {output_file} generated successfully!")
        except Exception as e:
            messagebox.showerror("Error", str(e))

root = tk.Tk()
app = App(root)
root.mainloop()

